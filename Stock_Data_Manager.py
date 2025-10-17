import os
import re
import yfinance as yf
import pandas as pd
import tkinter as tk
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border
from openpyxl.utils.dataframe import dataframe_to_rows
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


"""
period_options = [
    "1d",    # Last 1 day
    "5d",    # Last 5 days
    "1mo",   # Last 1 month
    "3mo",   # Last 3 months
    "6mo",   # Last 6 months
    "1y",    # Last 1 year
    "2y",    # Last 2 years
    "5y",    # Last 5 years
    "10y",   # Last 10 years
    "ytd",   # Year to date
    "max"    # All available data
]
interval_options = [
    "1m",    # 1 minute (only for last 7 days)
    "2m",    # 2 minutes
    "5m",    # 5 minutes
    "15m",   # 15 minutes
    "30m",   # 30 minutes
    "60m",   # 1 hour
    "90m",   # 1.5 hours
    "1d",    # Daily
    "5d",    # 5 days
    "1wk",   # Weekly
    "1mo",   # Monthly
    "3mo"    # Quarterly
]
"""

# Stock symbol
stock = 'TSLA'

### DATA OPTIONS ###
period = "6mo"
interval = "1d"

# Time range mode
use_custom_date = False
start_date = "2024-01-01"
end_date = "2024-06-01"
custom_interval = "1d"

# Visualization
show_chart = True

# Update old Excel files automatically
auto_update = True

if auto_update:
    folder_path = "./stocks"

    # Regex pattern: Stock-Interval-Period.xlsx
    pattern = re.compile(r"^(?P<stock>\w+)-(?P<interval>\w+)-(?P<period>\w+)\.xlsx$")

    # List files in the folder
    files = os.listdir(folder_path)

    for file in files:
        match = pattern.match(file)
        if match:
            current_stock_name = match.group("stock")
            current_interval = match.group("interval")
            current_period = match.group("period")

            # Full file path
            file_path = os.path.join(folder_path, file)

            ticker_data = yf.Ticker(current_stock_name)
            df = ticker_data.history(period=current_period, interval=current_interval)
            df.index = df.index.tz_localize(None)

            # Prepare OHLC + Volume data
            ohlc = df[["Open", "High", "Low", "Close", "Volume"]].copy()
            ohlc[["Open", "High", "Low", "Close"]] = ohlc[["Open", "High", "Low", "Close"]].round(2)
            ohlc["Date"] = df.index.strftime("%d.%m.%Y")
            ohlc["Time"] = "00:00:00"
            ohlc = ohlc[["Date", "Time", "Open", "High", "Low", "Close", "Volume"]]

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active

            # Styling
            font = Font(name='Calibri', size=11)
            align_right = Alignment(horizontal='right', vertical='top')
            align_left = Alignment(horizontal='left', vertical='top')
            border = Border()

            # Write DataFrame to Excel
            for i, row in enumerate(dataframe_to_rows(ohlc, index=False, header=True)):
                ws.append(row)
                for j, cell in enumerate(ws[i + 1]):  # +1 because openpyxl rows start from 1
                    cell.font = font
                    cell.border = border
                    if i == 0:
                        cell.alignment = align_left  # Header
                    else:
                        cell.alignment = align_right  # Data

            # Adjust column widths
            column_widths = {
                'A': 12,  # Date
                'B': 10,  # Time
                'C': 8,   # Open
                'D': 8,   # High
                'E': 8,   # Low
                'F': 8,   # Close
                'G': 12,  # Volume
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Save Excel file
            folder = "stocks"
            if not os.path.exists(folder):
                os.makedirs(folder)

            filename = f"{ticker_data.ticker}-{current_interval}-{current_period}.xlsx"
            filepath = os.path.join(folder, filename)

            # Remove existing file if present
            if os.path.exists(filepath):
                os.remove(filepath)

            wb.save(filepath)

# Fetch new data
target_stock = yf.Ticker(stock)
if use_custom_date:
    df = target_stock.history(start=start_date, end=end_date, interval=custom_interval)
else:
    df = target_stock.history(period=period, interval=interval)
df.index = df.index.tz_localize(None)

# Prepare OHLC + Volume data
ohlc = df[["Open", "High", "Low", "Close", "Volume"]].copy()
ohlc[["Open", "High", "Low", "Close"]] = ohlc[["Open", "High", "Low", "Close"]].round(2)
ohlc["Date"] = df.index.strftime("%d.%m.%Y")
ohlc["Time"] = "00:00:00"
ohlc = ohlc[["Date", "Time", "Open", "High", "Low", "Close", "Volume"]]

# Create Excel workbook
wb = Workbook()
ws = wb.active

# Styling
font = Font(name='Calibri', size=11)
align_right = Alignment(horizontal='right', vertical='top')
align_left = Alignment(horizontal='left', vertical='top')
border = Border()

# Write DataFrame to Excel
for i, row in enumerate(dataframe_to_rows(ohlc, index=False, header=True)):
    ws.append(row)
    for j, cell in enumerate(ws[i + 1]):
        cell.font = font
        cell.border = border
        if i == 0:
            cell.alignment = align_left
        else:
            cell.alignment = align_right

# Adjust column widths
column_widths = {
    'A': 12,
    'B': 10,
    'C': 8,
    'D': 8,
    'E': 8,
    'F': 8,
    'G': 12,
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save Excel file
folder = "stocks"
if not os.path.exists(folder):
    os.makedirs(folder)

if not use_custom_date:
    filename = f"{target_stock.ticker}-{interval}-{period}.xlsx"
else:
    filename = f"{target_stock.ticker}-{start_date}-{end_date}-{custom_interval}.xlsx"

filepath = os.path.join(folder, filename)

# Remove existing file if already exists
if os.path.exists(filepath):
    os.remove(filepath)

wb.save(filepath)
print(f"#### File saved successfully ####")

# Show chart
if show_chart:
    root = tk.Tk()
    root.title(f"{stock} Stock Chart ({period} - {interval})")
    root.geometry("1000x600")

    # Matplotlib figure
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(df.index, df['Close'], label='Closing Price', color='blue')
    ax.set_title(f'{stock} Stock - Closing Prices')
    ax.set_xlabel('Date')
    ax.set_ylabel('Price ($)')
    ax.grid(True)
    ax.legend()

    # Tkinter Canvas
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    root.mainloop()
